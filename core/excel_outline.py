"""Utilities for reading and writing Excel outline metadata."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


@dataclass
class OutlineNode:
    """Representation of a row/column outline group."""

    axis: str
    sheet: str
    level: int
    start: int
    end: int
    collapsed: bool
    children: List["OutlineNode"] = field(default_factory=list)

    def as_dict(self) -> Dict[str, object]:
        """Return a serialisable representation of the node."""

        return {
            "axis": self.axis,
            "sheet": self.sheet,
            "level": self.level,
            "start": self.start,
            "end": self.end,
            "collapsed": self.collapsed,
            "children": [child.as_dict() for child in self.children],
        }


def _should_parse(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def read_outline_levels(path: str) -> Dict[str, Dict[str, Dict[int, dict]]]:
    """Read outline metadata for all sheets in the workbook located at ``path``."""

    file_path = Path(path)
    if not file_path.exists() or not _should_parse(file_path):
        return {}

    workbook = load_workbook(path, data_only=True, read_only=False, keep_vba=True)
    outline: Dict[str, Dict[str, Dict[int, dict]]] = {}
    for ws in workbook.worksheets:
        row_map: Dict[int, dict] = {}
        col_map: Dict[int, dict] = {}

        for idx, dim in ws.row_dimensions.items():
            if not isinstance(idx, int):
                try:
                    idx = int(idx)
                except (TypeError, ValueError):  # pragma: no cover - defensive
                    continue
            level = int(getattr(dim, "outlineLevel", 0) or 0)
            hidden = bool(getattr(dim, "hidden", False))
            if level > 0 or hidden:
                row_map[int(idx)] = {"level": level, "hidden": hidden}

        for key, dim in ws.column_dimensions.items():
            if not key:
                continue
            try:
                col_idx = column_index_from_string(key)
            except ValueError:  # pragma: no cover - defensive
                continue
            level = int(getattr(dim, "outlineLevel", 0) or 0)
            hidden = bool(getattr(dim, "hidden", False))
            if level > 0 or hidden:
                col_map[int(col_idx)] = {"level": level, "hidden": hidden}

        outline[ws.title] = {"rows": row_map, "cols": col_map}
    return outline


def build_outline_nodes(level_map: Dict[int, dict], axis: str, sheet: str) -> List[OutlineNode]:
    """Build a nested outline tree from a mapping of indexes to outline metadata."""

    if not level_map:
        return []

    items = sorted((int(idx), meta) for idx, meta in level_map.items())
    stack: List[Dict[str, object]] = []
    completed: List[OutlineNode] = []

    def close_nodes(target_level: int, current_index: int) -> None:
        while stack and int(stack[-1]["level"]) > target_level:
            info = stack.pop()
            start = int(info["start"])
            end = current_index - 1
            if end < start:
                end = start
            node = OutlineNode(
                axis=axis,
                sheet=sheet,
                level=int(info["level"]),
                start=start,
                end=end,
                collapsed=bool(info["collapsed"]),
                children=list(info["children"]),
            )
            if stack:
                stack[-1]["children"].append(node)
            else:
                completed.append(node)

    last_index = items[-1][0]
    for index, meta in items:
        level = int(meta.get("level", 0) or 0)
        hidden = bool(meta.get("hidden", False))
        # close nodes if level decreased
        close_nodes(level, index)
        if level <= 0:
            continue
        # ensure stack has entries up to current level
        while len(stack) < level:
            stack.append(
                {
                    "level": len(stack) + 1,
                    "start": index,
                    "collapsed": False,
                    "children": [],
                }
            )
        if hidden:
            for info in stack:
                info["collapsed"] = bool(info["collapsed"]) or hidden

    close_nodes(0, last_index + 1)
    return completed


def extract_outline_tree(path: str) -> Dict[str, Dict[str, List[OutlineNode]]]:
    """Return outline trees for all sheets present in ``path``."""

    levels = read_outline_levels(path)
    tree: Dict[str, Dict[str, List[OutlineNode]]] = {}
    for sheet, axes in levels.items():
        row_nodes = build_outline_nodes(axes.get("rows", {}), axis="row", sheet=sheet)
        col_nodes = build_outline_nodes(axes.get("cols", {}), axis="col", sheet=sheet)
        tree[sheet] = {"rows": row_nodes, "cols": col_nodes}
    return tree


def _apply_nodes(ws, nodes: Iterable[OutlineNode]) -> None:
    for node in nodes:
        if node.axis == "row":
            for row_idx in range(node.start, node.end + 1):
                dim = ws.row_dimensions[row_idx]
                current_level = int(getattr(dim, "outlineLevel", 0) or 0)
                dim.outlineLevel = max(current_level, node.level)
                if node.collapsed:
                    dim.hidden = True
        else:
            for col_idx in range(node.start, node.end + 1):
                letter = get_column_letter(col_idx)
                dim = ws.column_dimensions[letter]
                current_level = int(getattr(dim, "outlineLevel", 0) or 0)
                dim.outlineLevel = max(current_level, node.level)
                if node.collapsed:
                    dim.hidden = True
        if node.children:
            _apply_nodes(ws, node.children)


def apply_outline_to_openpyxl(ws, nodes: List[OutlineNode]) -> None:
    """Apply outline metadata to an openpyxl worksheet."""

    if not nodes:
        return
    _apply_nodes(ws, nodes)
